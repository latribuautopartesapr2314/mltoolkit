import re
import unicodedata
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN APP
# ============================================================

st.set_page_config(
    page_title="ML Toolkit",
    page_icon="📦",
    layout="wide"
)


# ============================================================
# FUNCIONES COMUNES
# ============================================================

def normalizar_texto(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\s+", " ", texto)

    return texto.lower()


def es_vacio(valor):
    try:
        if pd.isna(valor):
            return True
    except Exception:
        pass

    return valor is None or str(valor).strip() == ""


def resolver_columna_por_nombre(columnas, nombre_buscado):
    buscado = normalizar_texto(nombre_buscado)

    for col in columnas:
        if normalizar_texto(col) == buscado:
            return col

    return None


def detectar_columna(columnas, posibles, obligatorio=True, descripcion="columna"):
    columnas_lista = list(columnas)

    for posible in posibles:
        real = resolver_columna_por_nombre(columnas_lista, posible)
        if real is not None:
            return real

    if obligatorio:
        raise ValueError(
            f"No se encontró {descripcion}. Columnas disponibles: "
            + ", ".join(map(str, columnas_lista))
        )

    return None


def convertir_precio_a_numero(valor):
    """
    Convierte precios tipo:
    - 67285
    - 67.285,00 ARS
    - $ 67.285,00
    - 67285.00
    a número.
    """

    if pd.isna(valor):
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if texto == "":
        return None

    texto = re.sub(r"[^\d,.\-]", "", texto)

    if texto == "":
        return None

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "")
            texto = texto.replace(",", ".")
        else:
            texto = texto.replace(",", "")

    elif "," in texto and "." not in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except Exception:
        return None


def convertir_stock(valor):
    numero = convertir_precio_a_numero(valor)

    if numero is None:
        return None

    try:
        return int(float(numero))
    except Exception:
        return None


def formatear_numero(valor):
    if valor is None:
        return None

    try:
        valor = float(valor)
    except Exception:
        return None

    if valor.is_integer():
        return int(valor)

    return round(valor, 2)


def normalizar_sku(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip().upper()

    if texto.endswith(".0"):
        texto = texto[:-2]

    return texto


def obtener_headers_ws(ws):
    headers = {}

    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    return headers


def obtener_o_crear_columna(ws, headers, nombre_columna):
    if nombre_columna in headers:
        return headers[nombre_columna]

    nueva_col = ws.max_column + 1
    ws.cell(row=1, column=nueva_col).value = nombre_columna
    headers[nombre_columna] = nueva_col

    return nueva_col


def buscar_hoja(wb, nombre_buscado):
    buscado = normalizar_texto(nombre_buscado)

    for hoja in wb.sheetnames:
        if normalizar_texto(hoja) == buscado:
            return hoja

    return None


def agregar_tabla(ws, titulo, encabezados, registros, fila_vacia):
    ws.append([])
    ws.append([titulo])
    ws.append(encabezados)

    if registros:
        for registro in registros:
            ws.append([registro.get(col, "") for col in encabezados])
    else:
        ws.append(fila_vacia)


def ajustar_anchos(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


# ============================================================
# HERRAMIENTA 1 — ACTUALIZAR INTEGRALY
# ============================================================

NOMBRE_HOJA_CONTROL_INTEGRALY = "CONTROL_CRUCE_INTEGRALY"

SUMA_FIJA_PRECIO = 12_000
UMBRAL_PRECIO_INTEGRALY = 1_500_000

ESTADO_ACTIVA = "Activa"
ESTADO_PAUSADA = "Pausada"

POSIBLES_SKU = [
    "sku", "SKU", "Sku",
    "item_code", "ITEM_CODE", "Item Code",
    "codigo", "Código", "CODIGO"
]

POSIBLES_STOCK_ORIGEN = [
    "AStk", "astk", "ASTK",
    "stock", "Stock", "STOCK"
]

POSIBLES_STOCK_DESTINO = [
    "cantidad", "Cantidad", "CANTIDAD",
    "stock", "Stock", "STOCK"
]

POSIBLES_PRECIO_DESTINO = [
    "precio", "Precio", "PRECIO",
    "price", "Price", "PRICE"
]

POSIBLES_ESTADO_DESTINO = [
    "estado", "Estado", "ESTADO",
    "status", "Status", "STATUS"
]

POSIBLES_CUOTAS = [
    "cuotas", "Cuotas", "CUOTAS"
]

POSIBLES_MLA = [
    "mla", "MLA", "Mla",
    "id", "ID",
    "publicacion", "Publicacion", "PUBLICACION"
]

COLUMNAS_PRECIO_ORIGEN = {
    "sin_cuotas": "Precio ML Clasica",
    "3_cuotas": "Precio ML Premium",
    "6_cuotas": "Precio ML Premium 6c",
    "9_cuotas": "Precio ML Premium 9c",
    "12_cuotas": "Precio ML Premium 12c",
}


def clasificar_cuotas(valor):
    if pd.isna(valor):
        return "sin_cuotas"

    texto = str(valor).strip().lower()

    if texto == "":
        return "sin_cuotas"

    texto = (
        texto
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )

    if "no agregar" in texto:
        return "sin_cuotas"

    if "sin cuota" in texto:
        return "sin_cuotas"

    if re.search(r"\b12\b", texto):
        return "12_cuotas"

    if re.search(r"\b9\b", texto):
        return "9_cuotas"

    if re.search(r"\b6\b", texto):
        return "6_cuotas"

    if re.search(r"\b3\b", texto):
        return "3_cuotas"

    return None


def determinar_estado_por_stock(stock_final):
    if stock_final is None:
        return None, "No se pudo determinar stock"

    try:
        stock_num = int(float(stock_final))
    except Exception:
        return None, "Stock inválido"

    if stock_num >= 1:
        return ESTADO_ACTIVA, "Precio actual Integraly menor o igual a 1.500.000 y stock >= 1"

    return ESTADO_PAUSADA, "Precio actual Integraly menor o igual a 1.500.000 y stock <= 0"


def armar_base_global_precios(actualizacion_bytes):
    xls = pd.ExcelFile(BytesIO(actualizacion_bytes))

    filas_origen = []
    control_hojas_ignoradas = []
    control_precios_invalidos = []
    control_stock_invalido = []

    for hoja in xls.sheet_names:

        if str(hoja).upper().startswith("CONTROL"):
            continue

        df = pd.read_excel(
            BytesIO(actualizacion_bytes),
            sheet_name=hoja,
            dtype=str
        )

        if df.empty:
            control_hojas_ignoradas.append({
                "Hoja": hoja,
                "Motivo": "Hoja vacía"
            })
            continue

        columnas = list(df.columns)

        try:
            col_sku = detectar_columna(
                columnas,
                POSIBLES_SKU,
                obligatorio=True,
                descripcion="columna SKU origen"
            )

            col_stock = detectar_columna(
                columnas,
                POSIBLES_STOCK_ORIGEN,
                obligatorio=True,
                descripcion="columna stock origen"
            )

        except Exception as e:
            control_hojas_ignoradas.append({
                "Hoja": hoja,
                "Motivo": str(e)
            })
            continue

        columnas_precio_reales = {}
        faltan_precios = []

        for clave, col_esperada in COLUMNAS_PRECIO_ORIGEN.items():
            col_real = resolver_columna_por_nombre(columnas, col_esperada)

            if col_real is None:
                faltan_precios.append(col_esperada)
            else:
                columnas_precio_reales[clave] = col_real

        if faltan_precios:
            control_hojas_ignoradas.append({
                "Hoja": hoja,
                "Motivo": "Faltan columnas de precio: " + ", ".join(faltan_precios)
            })
            continue

        df["_SKU_KEY"] = df[col_sku].apply(normalizar_sku)
        df = df[df["_SKU_KEY"] != ""].copy()

        for idx, fila in df.iterrows():

            sku_key = fila["_SKU_KEY"]
            stock_final = convertir_stock(fila[col_stock])

            if stock_final is None:
                control_stock_invalido.append({
                    "Hoja origen": hoja,
                    "Fila origen": idx + 2,
                    "SKU": sku_key,
                    "Valor AStk": fila[col_stock],
                })

            registro = {
                "_SKU_KEY": sku_key,
                "_HOJA_ORIGEN": hoja,
                "_FILA_ORIGEN": idx + 2,
                "_ASTK": stock_final,
            }

            for clave_precio, col_precio in columnas_precio_reales.items():

                precio_base = convertir_precio_a_numero(fila[col_precio])

                if precio_base is not None:
                    precio_final = formatear_numero(precio_base + SUMA_FIJA_PRECIO)
                else:
                    precio_final = None

                registro[clave_precio] = precio_final

                if precio_final is None:
                    control_precios_invalidos.append({
                        "Hoja origen": hoja,
                        "Fila origen": idx + 2,
                        "SKU": sku_key,
                        "Columna precio": col_precio,
                        "Valor original": fila[col_precio],
                    })

            filas_origen.append(registro)

    if not filas_origen:
        raise ValueError(
            "No se pudo armar la base de actualización. "
            "Revisá que el archivo tenga SKU, AStk y columnas de precio sin _2."
        )

    df_origen = pd.DataFrame(filas_origen)

    duplicados = df_origen[df_origen["_SKU_KEY"].duplicated(keep=False)].copy()

    df_unico = df_origen.drop_duplicates(
        subset="_SKU_KEY",
        keep="first"
    ).copy()

    datos_por_sku = df_unico.set_index("_SKU_KEY").to_dict(orient="index")

    return (
        datos_por_sku,
        duplicados,
        control_hojas_ignoradas,
        control_precios_invalidos,
        control_stock_invalido
    )


def procesar_integraly(integraly_bytes, actualizacion_bytes):
    (
        datos_por_sku,
        duplicados,
        hojas_ignoradas,
        precios_invalidos,
        stock_invalido
    ) = armar_base_global_precios(actualizacion_bytes)

    wb = load_workbook(BytesIO(integraly_bytes))

    if NOMBRE_HOJA_CONTROL_INTEGRALY in wb.sheetnames:
        del wb[NOMBRE_HOJA_CONTROL_INTEGRALY]

    control_resumen = []
    control_sku_sin_match = []
    control_cuotas_no_reconocidas = []
    control_precio_no_actualizado = []
    control_stock_no_actualizado = []
    control_estado_actualizado = []
    control_estado_no_actualizado = []
    control_mayor_umbral_pausada = []
    control_bloqueados_por_umbral_integraly = []

    for ws in wb.worksheets:

        if ws.title.upper().startswith("CONTROL"):
            continue

        headers = obtener_headers_ws(ws)

        col_sku_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_SKU,
            obligatorio=False,
            descripcion="columna SKU destino"
        )

        if col_sku_nombre is None:
            control_resumen.append({
                "Hoja destino": ws.title,
                "Estado hoja": "Ignorada",
                "Motivo": "No se encontró columna SKU",
                "Filas procesadas": 0,
                "Filas actualizadas con precio": 0,
                "Filas actualizadas con stock": 0,
                "SKU sin match": 0,
                "Cuotas no reconocidas": 0,
                "Precio no actualizado": 0,
                "Stock no actualizado": 0,
                "Estado actualizado": 0,
                "Estado no actualizado": 0,
                "Pausadas por precio Integraly mayor a 1.500.000": 0,
                "Bloqueadas sin modificar precio/stock": 0,
            })
            continue

        col_mla_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_MLA,
            obligatorio=False,
            descripcion="columna MLA"
        )

        col_cuotas_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_CUOTAS,
            obligatorio=False,
            descripcion="columna cuotas"
        )

        col_precio_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_PRECIO_DESTINO,
            obligatorio=False,
            descripcion="columna precio destino"
        )

        col_stock_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_STOCK_DESTINO,
            obligatorio=False,
            descripcion="columna stock destino"
        )

        col_estado_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_ESTADO_DESTINO,
            obligatorio=False,
            descripcion="columna estado destino"
        )

        if col_precio_nombre is None:
            col_precio_nombre = "precio"
            obtener_o_crear_columna(ws, headers, col_precio_nombre)

        if col_stock_nombre is None:
            col_stock_nombre = "cantidad"
            obtener_o_crear_columna(ws, headers, col_stock_nombre)

        if col_estado_nombre is None:
            col_estado_nombre = "estado"
            obtener_o_crear_columna(ws, headers, col_estado_nombre)

        col_sku = headers[col_sku_nombre]
        col_precio = headers[col_precio_nombre]
        col_stock = headers[col_stock_nombre]
        col_estado = headers[col_estado_nombre]

        col_mla = headers[col_mla_nombre] if col_mla_nombre else None
        col_cuotas = headers[col_cuotas_nombre] if col_cuotas_nombre else None

        filas_procesadas = 0
        filas_actualizadas_precio = 0
        filas_actualizadas_stock = 0
        sku_sin_match = 0
        cuotas_no_reconocidas = 0
        precio_no_actualizado = 0
        stock_no_actualizado = 0
        estado_actualizado = 0
        estado_no_actualizado = 0
        pausadas_umbral_integraly = 0
        bloqueadas_umbral_integraly = 0

        for row in range(2, ws.max_row + 1):

            sku_original = ws.cell(row=row, column=col_sku).value
            sku_key = normalizar_sku(sku_original)

            if sku_key == "":
                continue

            filas_procesadas += 1

            mla = ws.cell(row=row, column=col_mla).value if col_mla else ""
            cuotas_original = ws.cell(row=row, column=col_cuotas).value if col_cuotas else ""
            estado_anterior = ws.cell(row=row, column=col_estado).value
            precio_anterior = ws.cell(row=row, column=col_precio).value
            stock_anterior = ws.cell(row=row, column=col_stock).value

            precio_integraly_actual_num = convertir_precio_a_numero(precio_anterior)

            if sku_key not in datos_por_sku:
                sku_sin_match += 1
                control_sku_sin_match.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Cuotas": cuotas_original,
                    "Precio actual Integraly": precio_anterior,
                    "Estado anterior": estado_anterior,
                })
                continue

            datos = datos_por_sku[sku_key]
            stock_final = datos.get("_ASTK")

            clave_precio = clasificar_cuotas(cuotas_original)

            if clave_precio is None:
                cuotas_no_reconocidas += 1
                estado_no_actualizado += 1

                control_cuotas_no_reconocidas.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Cuotas": cuotas_original,
                    "Precio actual Integraly": precio_anterior,
                })

                control_estado_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": "",
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Motivo": "Cuotas no reconocidas, no se pudo determinar precio",
                })

                continue

            precio_final = datos.get(clave_precio)

            if precio_final is None:
                precio_no_actualizado += 1
                estado_no_actualizado += 1

                control_precio_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Cuotas": cuotas_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio requerido": COLUMNAS_PRECIO_ORIGEN[clave_precio],
                    "Hoja origen": datos.get("_HOJA_ORIGEN"),
                    "Fila origen": datos.get("_FILA_ORIGEN"),
                })

                control_estado_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Motivo": "Precio final Global + 12000 inválido o vacío",
                })

                continue

            # ====================================================
            # REGLA CRÍTICA:
            # SI EL PRECIO ACTUAL DE INTEGRALY ES MAYOR A 1.500.000:
            # - NO MODIFICA PRECIO
            # - NO MODIFICA STOCK
            # - SOLO MODIFICA ESTADO A PAUSADA
            # ====================================================

            if (
                precio_integraly_actual_num is not None
                and precio_integraly_actual_num > UMBRAL_PRECIO_INTEGRALY
            ):

                ws.cell(row=row, column=col_estado).value = ESTADO_PAUSADA

                estado_actualizado += 1
                pausadas_umbral_integraly += 1
                bloqueadas_umbral_integraly += 1

                registro_bloqueado = {
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Cuotas": cuotas_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio actual Integraly numérico": precio_integraly_actual_num,
                    "Stock actual Integraly": stock_anterior,
                    "Precio calculado Global + 12000": precio_final,
                    "Stock Global AStk": stock_final,
                    "Estado anterior": estado_anterior,
                    "Estado final": ESTADO_PAUSADA,
                    "Motivo": "Precio actual de Integraly mayor a 1.500.000: no se modifica precio ni stock, solo se pausa",
                }

                control_bloqueados_por_umbral_integraly.append(registro_bloqueado)

                control_mayor_umbral_pausada.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Estado final": ESTADO_PAUSADA,
                    "Motivo": "Precio actual de Integraly mayor a 1.500.000: no se modifica precio ni stock, solo se pausa",
                })

                control_estado_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Estado final": ESTADO_PAUSADA,
                    "Motivo": "Precio actual de Integraly mayor a 1.500.000: no se modifica precio ni stock, solo se pausa",
                })

                continue

            # ====================================================
            # SI EL PRECIO ACTUAL DE INTEGRALY ES MENOR O IGUAL
            # A 1.500.000, O NO SE PUDO LEER:
            # - MODIFICA PRECIO DESDE GLOBAL + 12000
            # - MODIFICA STOCK DESDE ASTK
            # - MODIFICA ESTADO SEGÚN STOCK
            # ====================================================

            ws.cell(row=row, column=col_precio).value = precio_final
            filas_actualizadas_precio += 1

            if stock_final is not None:
                ws.cell(row=row, column=col_stock).value = stock_final
                filas_actualizadas_stock += 1
            else:
                stock_no_actualizado += 1
                control_stock_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Valor AStk": stock_final,
                    "Hoja origen": datos.get("_HOJA_ORIGEN"),
                    "Fila origen": datos.get("_FILA_ORIGEN"),
                })

            estado_final, motivo_estado = determinar_estado_por_stock(stock_final)

            if estado_final is not None:
                ws.cell(row=row, column=col_estado).value = estado_final
                estado_actualizado += 1

                control_estado_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Estado final": estado_final,
                    "Motivo": motivo_estado,
                })
            else:
                estado_no_actualizado += 1

                control_estado_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Motivo": motivo_estado,
                })

        control_resumen.append({
            "Hoja destino": ws.title,
            "Estado hoja": "Procesada",
            "Motivo": "",
            "Filas procesadas": filas_procesadas,
            "Filas actualizadas con precio": filas_actualizadas_precio,
            "Filas actualizadas con stock": filas_actualizadas_stock,
            "SKU sin match": sku_sin_match,
            "Cuotas no reconocidas": cuotas_no_reconocidas,
            "Precio no actualizado": precio_no_actualizado,
            "Stock no actualizado": stock_no_actualizado,
            "Estado actualizado": estado_actualizado,
            "Estado no actualizado": estado_no_actualizado,
            "Pausadas por precio Integraly mayor a 1.500.000": pausadas_umbral_integraly,
            "Bloqueadas sin modificar precio/stock": bloqueadas_umbral_integraly,
        })

    ws_control = wb.create_sheet(NOMBRE_HOJA_CONTROL_INTEGRALY)

    ws_control.append(["Fecha proceso", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_control.append(["Suma fija aplicada a cada precio Global", SUMA_FIJA_PRECIO])
    ws_control.append(["Umbral precio actual Integraly", UMBRAL_PRECIO_INTEGRALY])
    ws_control.append(["Estado activa", ESTADO_ACTIVA])
    ws_control.append(["Estado pausada", ESTADO_PAUSADA])
    ws_control.append([])

    ws_control.append(["MAPEO DE CUOTAS"])
    ws_control.append(["Condición en Integraly", "Columna usada en Global"])
    ws_control.append(["Sin cuotas / No Agregar Cuotas / vacío", "Precio ML Clasica + 12000"])
    ws_control.append(["3 cuotas", "Precio ML Premium + 12000"])
    ws_control.append(["6 cuotas", "Precio ML Premium 6c + 12000"])
    ws_control.append(["9 cuotas", "Precio ML Premium 9c + 12000"])
    ws_control.append(["12 cuotas", "Precio ML Premium 12c + 12000"])
    ws_control.append(["Stock", "AStk"])
    ws_control.append([])

    ws_control.append(["REGLA CRÍTICA PRECIO ACTUAL INTEGRALY"])
    ws_control.append(["Condición", "Acción"])
    ws_control.append(["Precio actual Integraly > 1.500.000", "NO modifica precio, NO modifica stock, SOLO estado = Pausada"])
    ws_control.append(["Precio actual Integraly <= 1.500.000 y stock Global >= 1", "Modifica precio, modifica stock, estado = Activa"])
    ws_control.append(["Precio actual Integraly <= 1.500.000 y stock Global <= 0", "Modifica precio, modifica stock, estado = Pausada"])
    ws_control.append(["Precio actual Integraly vacío/no legible", "Modifica precio y stock si hay match; estado según stock"])
    ws_control.append([])

    ws_control.append(["RESUMEN"])

    resumen_cols = [
        "Hoja destino",
        "Estado hoja",
        "Motivo",
        "Filas procesadas",
        "Filas actualizadas con precio",
        "Filas actualizadas con stock",
        "SKU sin match",
        "Cuotas no reconocidas",
        "Precio no actualizado",
        "Stock no actualizado",
        "Estado actualizado",
        "Estado no actualizado",
        "Pausadas por precio Integraly mayor a 1.500.000",
        "Bloqueadas sin modificar precio/stock",
    ]

    ws_control.append(resumen_cols)

    for r in control_resumen:
        ws_control.append([r.get(c, "") for c in resumen_cols])

    agregar_tabla(
        ws_control,
        "BLOQUEADAS POR PRECIO ACTUAL INTEGRALY MAYOR A 1.500.000 - NO SE MODIFICÓ PRECIO NI STOCK",
        [
            "Hoja destino",
            "Fila destino",
            "MLA",
            "SKU",
            "Cuotas",
            "Precio actual Integraly",
            "Precio actual Integraly numérico",
            "Stock actual Integraly",
            "Precio calculado Global + 12000",
            "Stock Global AStk",
            "Estado anterior",
            "Estado final",
            "Motivo",
        ],
        control_bloqueados_por_umbral_integraly,
        ["Sin bloqueadas por precio Integraly mayor a 1.500.000", "", "", "", "", "", "", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "SKU SIN MATCH",
        ["Hoja destino", "Fila destino", "MLA", "SKU", "Cuotas", "Precio actual Integraly", "Estado anterior"],
        control_sku_sin_match,
        ["Sin SKU sin match", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "CUOTAS NO RECONOCIDAS",
        ["Hoja destino", "Fila destino", "MLA", "SKU", "Cuotas", "Precio actual Integraly"],
        control_cuotas_no_reconocidas,
        ["Sin cuotas no reconocidas", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "PRECIO NO ACTUALIZADO",
        ["Hoja destino", "Fila destino", "MLA", "SKU", "Cuotas", "Precio actual Integraly", "Precio requerido", "Hoja origen", "Fila origen"],
        control_precio_no_actualizado,
        ["Sin precios pendientes", "", "", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "STOCK NO ACTUALIZADO",
        ["Hoja destino", "Fila destino", "MLA", "SKU", "Valor AStk", "Hoja origen", "Fila origen"],
        control_stock_no_actualizado,
        ["Sin stocks pendientes", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "ESTADO ACTUALIZADO",
        [
            "Hoja destino",
            "Fila destino",
            "MLA",
            "SKU",
            "Precio actual Integraly",
            "Precio final Global + 12000",
            "Stock final",
            "Estado anterior",
            "Estado final",
            "Motivo",
        ],
        control_estado_actualizado,
        ["Sin estados actualizados", "", "", "", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "ESTADO NO ACTUALIZADO",
        [
            "Hoja destino",
            "Fila destino",
            "MLA",
            "SKU",
            "Precio actual Integraly",
            "Precio final Global + 12000",
            "Stock final",
            "Estado anterior",
            "Motivo",
        ],
        control_estado_no_actualizado,
        ["Sin estados pendientes", "", "", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "PAUSADAS POR PRECIO ACTUAL INTEGRALY MAYOR A 1.500.000",
        [
            "Hoja destino",
            "Fila destino",
            "MLA",
            "SKU",
            "Precio actual Integraly",
            "Precio final Global + 12000",
            "Stock final",
            "Estado anterior",
            "Estado final",
            "Motivo",
        ],
        control_mayor_umbral_pausada,
        ["Sin publicaciones mayores a 1.500.000 pausadas", "", "", "", "", "", "", "", "", ""]
    )

    ws_control.append([])
    ws_control.append(["SKU DUPLICADOS EN GLOBAL / ACTUALIZACIÓN PRECIO"])
    ws_control.append(["SKU", "Cantidad apariciones", "Hojas origen"])

    if not duplicados.empty:
        dup_resumen = (
            duplicados
            .groupby("_SKU_KEY")
            .agg(
                cantidad_apariciones=("_SKU_KEY", "size"),
                hojas_origen=("_HOJA_ORIGEN", lambda x: " | ".join(sorted(set(map(str, x)))))
            )
            .reset_index()
        )

        for _, r in dup_resumen.iterrows():
            ws_control.append([
                r["_SKU_KEY"],
                int(r["cantidad_apariciones"]),
                r["hojas_origen"]
            ])
    else:
        ws_control.append(["Sin duplicados", 0, ""])

    agregar_tabla(
        ws_control,
        "HOJAS IGNORADAS EN GLOBAL / ACTUALIZACIÓN PRECIO",
        ["Hoja", "Motivo"],
        hojas_ignoradas,
        ["Sin hojas ignoradas", ""]
    )

    agregar_tabla(
        ws_control,
        "PRECIOS INVÁLIDOS EN GLOBAL / ACTUALIZACIÓN PRECIO",
        ["Hoja origen", "Fila origen", "SKU", "Columna precio", "Valor original"],
        precios_invalidos,
        ["Sin precios inválidos", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "STOCK INVÁLIDO EN GLOBAL / ACTUALIZACIÓN PRECIO",
        ["Hoja origen", "Fila origen", "SKU", "Valor AStk"],
        stock_invalido,
        ["Sin stocks inválidos", "", "", ""]
    )

    ajustar_anchos(ws_control)

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)

    resumen = {
        "control_resumen": control_resumen,
        "sku_unicos_origen": len(datos_por_sku),
        "sku_duplicados_origen": duplicados["_SKU_KEY"].nunique() if not duplicados.empty else 0,
        "hojas_ignoradas_origen": len(hojas_ignoradas),
    }

    return salida, resumen


def mostrar_herramienta_integraly():
    st.title("Actualizar Integraly")
    st.caption("Actualiza precio, stock y estado por SKU.")

    st.warning(
        "La herramienta toma los precios del archivo Global, limpia ARS, convierte a número "
        "y suma $12.000 fijos antes de completar Integraly."
    )

    st.error(
        "Regla crítica: si el precio actual de Integraly supera $1.500.000, "
        "NO se modifica precio ni stock. Solo se fuerza el estado a Pausada."
    )

    st.info(
        "Mapeo: Clásica = sin cuotas | Premium = 3 cuotas | Premium 6c = 6 cuotas | "
        "Premium 9c = 9 cuotas | Premium 12c = 12 cuotas."
    )

    col1, col2 = st.columns(2)

    with col1:
        archivo_integraly = st.file_uploader(
            "Subí el archivo Integraly",
            type=["xlsx"],
            key="integraly_uploader"
        )

    with col2:
        archivo_actualizacion = st.file_uploader(
            "Subí el archivo Global / Actualización Precio",
            type=["xlsx"],
            key="actualizacion_uploader"
        )

    if st.button("Procesar Integraly", type="primary"):

        if archivo_integraly is None or archivo_actualizacion is None:
            st.error("Tenés que subir los 2 archivos para procesar.")

        else:
            try:
                with st.spinner("Procesando archivos..."):
                    salida, resumen = procesar_integraly(
                        archivo_integraly.getvalue(),
                        archivo_actualizacion.getvalue()
                    )

                st.success("Proceso finalizado correctamente.")

                st.subheader("Resumen")
                st.dataframe(
                    pd.DataFrame(resumen["control_resumen"]),
                    use_container_width=True
                )

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("SKU únicos origen", resumen["sku_unicos_origen"])
                c2.metric("SKU duplicados origen", resumen["sku_duplicados_origen"])
                c3.metric("Hojas ignoradas origen", resumen["hojas_ignoradas_origen"])
                c4.metric("Suma fija aplicada", f"${SUMA_FIJA_PRECIO:,.0f}")

                fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_salida = f"INTEGRALY_PRECIOS_STOCK_ESTADO_ACTUALIZADO_{fecha}.xlsx"

                st.download_button(
                    label="Descargar Excel Integraly actualizado",
                    data=salida,
                    file_name=nombre_salida,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error("El proceso falló.")
                st.exception(e)


# ============================================================
# HERRAMIENTA 2 — AGENTE PUBLICADOR
# ============================================================

NOMBRE_HOJA_CONFIG_PUBLICADOR = "CONFIG_PUBLICADOR"

HOJAS_AGENTE_BASE = [
    "BASE_COMPLETAR",
    "BASE_TITULOS",
    "BASE_DESCRIPCIONES",
    "BASE_IMAGENES",
]

LIMPIAR_FILAS_PUBLICAR_ANTES_DE_CARGAR = True


def leer_config_publicador(wb_agente):
    hoja_config = buscar_hoja(wb_agente, NOMBRE_HOJA_CONFIG_PUBLICADOR)

    if hoja_config is None:
        raise ValueError(
            f"No existe la hoja '{NOMBRE_HOJA_CONFIG_PUBLICADOR}'. "
            "Para que el sistema sea escalable, agregá esa hoja al AGENTE PUBLICADOR BASE."
        )

    ws = wb_agente[hoja_config]

    headers = {}

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(1, col).value

        if not es_vacio(valor):
            headers[normalizar_texto(valor)] = col

    columnas_requeridas = [
        "activo",
        "categoria",
        "hoja_publicar",
        "fila_modelo_base_completar",
    ]

    for requerida in columnas_requeridas:
        if requerida not in headers:
            raise ValueError(f"Falta la columna obligatoria '{requerida}' en CONFIG_PUBLICADOR.")

    config = []

    for row in range(2, ws.max_row + 1):
        activo = ws.cell(row, headers["activo"]).value
        categoria = ws.cell(row, headers["categoria"]).value
        hoja_publicar = ws.cell(row, headers["hoja_publicar"]).value
        fila_modelo = ws.cell(row, headers["fila_modelo_base_completar"]).value

        if es_vacio(categoria):
            continue

        if normalizar_texto(activo) not in ["si", "s", "yes", "true", "1"]:
            continue

        if es_vacio(hoja_publicar):
            hoja_publicar = categoria

        if es_vacio(fila_modelo):
            raise ValueError(f"La categoría '{categoria}' no tiene FILA_MODELO_BASE_COMPLETAR definida.")

        fila_encabezados = 3
        fila_inicio = 8

        if "fila_encabezados_publicar" in headers:
            valor = ws.cell(row, headers["fila_encabezados_publicar"]).value

            if not es_vacio(valor):
                fila_encabezados = int(valor)

        if "fila_inicio_publicar" in headers:
            valor = ws.cell(row, headers["fila_inicio_publicar"]).value

            if not es_vacio(valor):
                fila_inicio = int(valor)

        config.append({
            "categoria": str(categoria).strip(),
            "hoja_publicar": str(hoja_publicar).strip(),
            "fila_modelo": int(fila_modelo),
            "fila_encabezados": int(fila_encabezados),
            "fila_inicio": int(fila_inicio),
        })

    if not config:
        raise ValueError("CONFIG_PUBLICADOR no tiene categorías activas para procesar.")

    return config


def cargar_df_agente(agente_bytes):
    return pd.read_excel(
        BytesIO(agente_bytes),
        sheet_name=None,
        header=None,
        dtype=object,
        engine="openpyxl"
    )


def obtener_df(dfs, nombre_hoja):
    objetivo = normalizar_texto(nombre_hoja)

    for hoja, df in dfs.items():
        if normalizar_texto(hoja) == objetivo:
            return df

    raise ValueError(f"No se encontró la hoja '{nombre_hoja}' en el AGENTE.")


def buscar_fila_categoria(df, categoria):
    objetivo = normalizar_texto(categoria)

    for idx in range(len(df)):
        valor = df.iat[idx, 0]

        if normalizar_texto(valor) == objetivo:
            return idx

    return None


def extraer_titulos(df_titulos, categoria, categorias_validas):
    fila_categoria = buscar_fila_categoria(df_titulos, categoria)

    if fila_categoria is None:
        return []

    categorias_norm = {normalizar_texto(c) for c in categorias_validas}
    titulos = []

    for idx in range(fila_categoria + 1, len(df_titulos)):
        valor = df_titulos.iat[idx, 0]

        if es_vacio(valor):
            break

        texto = str(valor).strip()
        texto_norm = normalizar_texto(texto)

        if texto_norm in categorias_norm:
            break

        titulos.append(texto)

    return titulos


def es_marcador_descripcion(valor):
    valor_norm = normalizar_texto(valor)
    return valor_norm.startswith("descripcion")


def extraer_descripcion(df_descripciones, categoria, categorias_validas):
    categorias_norm = {normalizar_texto(c) for c in categorias_validas}

    fila_categoria = buscar_fila_categoria(df_descripciones, categoria)

    if fila_categoria is None:
        return ""

    lineas = []
    empezo = False
    encontro_marcador = False

    for idx in range(fila_categoria + 1, len(df_descripciones)):
        valor = df_descripciones.iat[idx, 0]

        if es_vacio(valor):
            if empezo:
                lineas.append("")
            continue

        texto = str(valor).strip()
        texto_norm = normalizar_texto(texto)

        if texto_norm in categorias_norm:
            break

        if es_marcador_descripcion(texto):
            if encontro_marcador and empezo:
                break

            encontro_marcador = True
            continue

        lineas.append(texto)
        empezo = True

    descripcion = "\n".join(lineas).strip()
    descripcion = re.sub(r"\n{3,}", "\n\n", descripcion)

    return descripcion


def extraer_imagenes(df_imagenes, categoria, categorias_validas):
    fila_categoria = buscar_fila_categoria(df_imagenes, categoria)

    if fila_categoria is None:
        return ""

    categorias_norm = {normalizar_texto(c) for c in categorias_validas}

    for idx in range(fila_categoria + 1, len(df_imagenes)):
        valor = df_imagenes.iat[idx, 0]

        if es_vacio(valor):
            continue

        texto = str(valor).strip()
        texto_norm = normalizar_texto(texto)

        if texto_norm in categorias_norm:
            break

        if "http" in texto_norm:
            return texto

    return ""


def buscar_columna_por_encabezado(ws, fila_encabezados, palabras_obligatorias):
    for col in range(1, ws.max_column + 1):
        encabezado = normalizar_texto(ws.cell(fila_encabezados, col).value)

        if all(palabra in encabezado for palabra in palabras_obligatorias):
            return col

    return None


def validar_columna(columna, nombre, hoja):
    if columna is None:
        raise ValueError(f"No se encontró la columna '{nombre}' en la hoja Publicar '{hoja}'.")


def columnas_que_no_deben_tocarse(ws, fila_encabezados):
    columnas = set()

    for col in range(1, ws.max_column + 1):
        encabezado = normalizar_texto(ws.cell(fila_encabezados, col).value)

        if "buybox_formula" in encabezado:
            columnas.add(col)

        if "hidden_pictures" in encabezado:
            columnas.add(col)

    return columnas


def limpiar_filas_publicar(ws, fila_inicio, columnas_saltar):
    for fila in range(fila_inicio, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if col in columnas_saltar:
                continue

            ws.cell(fila, col).value = None


def copiar_fila_modelo(ws_origen, ws_destino, fila_origen, fila_destino, columnas_saltar):
    max_col = min(ws_origen.max_column, ws_destino.max_column)

    for col in range(1, max_col + 1):
        if col in columnas_saltar:
            continue

        ws_destino.cell(fila_destino, col).value = ws_origen.cell(fila_origen, col).value


def contar_links_imagenes(texto):
    if es_vacio(texto):
        return 0

    partes = [x.strip() for x in str(texto).split(",")]

    return len([x for x in partes if x.startswith("http")])


def validar_estructura_agente(wb_agente):
    hojas_norm = {normalizar_texto(h) for h in wb_agente.sheetnames}
    faltantes = []

    for hoja in HOJAS_AGENTE_BASE:
        if normalizar_texto(hoja) not in hojas_norm:
            faltantes.append(hoja)

    if faltantes:
        raise ValueError(
            "El archivo AGENTE no tiene todas las hojas base requeridas: "
            + ", ".join(faltantes)
        )


def procesar_agente_publicador(agente_bytes, publicar_bytes):
    dfs_agente = cargar_df_agente(agente_bytes)

    wb_agente = load_workbook(BytesIO(agente_bytes), data_only=False)
    wb_publicar = load_workbook(BytesIO(publicar_bytes), data_only=False)

    validar_estructura_agente(wb_agente)

    config = leer_config_publicador(wb_agente)

    df_titulos = obtener_df(dfs_agente, "BASE_TITULOS")
    df_descripciones = obtener_df(dfs_agente, "BASE_DESCRIPCIONES")
    df_imagenes = obtener_df(dfs_agente, "BASE_IMAGENES")

    hoja_base_completar = buscar_hoja(wb_agente, "BASE_COMPLETAR")

    if hoja_base_completar is None:
        raise ValueError("No se encontró BASE_COMPLETAR en el AGENTE.")

    ws_base_completar = wb_agente[hoja_base_completar]

    categorias_validas = [item["categoria"] for item in config]

    resumen = []
    errores = []

    for item in config:
        categoria = item["categoria"]
        hoja_publicar_objetivo = item["hoja_publicar"]
        fila_modelo = item["fila_modelo"]
        fila_encabezados = item["fila_encabezados"]
        fila_inicio = item["fila_inicio"]

        hoja_publicar_real = buscar_hoja(wb_publicar, hoja_publicar_objetivo)

        if hoja_publicar_real is None:
            errores.append({
                "categoria": categoria,
                "error": f"No existe la hoja '{hoja_publicar_objetivo}' en Publicar."
            })
            continue

        ws_publicar = wb_publicar[hoja_publicar_real]

        if fila_modelo > ws_base_completar.max_row:
            errores.append({
                "categoria": categoria,
                "error": f"La fila modelo {fila_modelo} no existe en BASE_COMPLETAR."
            })
            continue

        titulos = extraer_titulos(df_titulos, categoria, categorias_validas)
        descripcion = extraer_descripcion(df_descripciones, categoria, categorias_validas)
        imagenes = extraer_imagenes(df_imagenes, categoria, categorias_validas)

        if len(titulos) == 0:
            errores.append({
                "categoria": categoria,
                "error": "No se encontraron títulos debajo de la categoría en BASE_TITULOS."
            })
            continue

        if es_vacio(descripcion):
            errores.append({
                "categoria": categoria,
                "error": "No se encontró descripción en BASE_DESCRIPCIONES."
            })
            continue

        if es_vacio(imagenes):
            errores.append({
                "categoria": categoria,
                "error": "No se encontraron links de imágenes en BASE_IMAGENES."
            })
            continue

        col_titulo = buscar_columna_por_encabezado(ws_publicar, fila_encabezados, ["titulo"])
        col_caracteres = buscar_columna_por_encabezado(ws_publicar, fila_encabezados, ["cantidad", "caracteres"])
        col_fotos = buscar_columna_por_encabezado(ws_publicar, fila_encabezados, ["fotos"])
        col_descripcion = buscar_columna_por_encabezado(ws_publicar, fila_encabezados, ["descripcion"])

        try:
            validar_columna(col_titulo, "Título", hoja_publicar_real)
            validar_columna(col_fotos, "Fotos", hoja_publicar_real)
            validar_columna(col_descripcion, "Descripción", hoja_publicar_real)
        except Exception as e:
            errores.append({
                "categoria": categoria,
                "error": str(e)
            })
            continue

        ultima_fila_necesaria = fila_inicio + len(titulos) - 1

        if ultima_fila_necesaria > ws_publicar.max_row:
            errores.append({
                "categoria": categoria,
                "error": (
                    f"La hoja '{hoja_publicar_real}' no tiene filas suficientes. "
                    f"Necesita hasta fila {ultima_fila_necesaria}, pero llega hasta {ws_publicar.max_row}."
                )
            })
            continue

        columnas_saltar = columnas_que_no_deben_tocarse(ws_publicar, fila_encabezados)

        if LIMPIAR_FILAS_PUBLICAR_ANTES_DE_CARGAR:
            limpiar_filas_publicar(ws_publicar, fila_inicio, columnas_saltar)

        for i, titulo in enumerate(titulos):
            fila_destino = fila_inicio + i

            copiar_fila_modelo(
                ws_origen=ws_base_completar,
                ws_destino=ws_publicar,
                fila_origen=fila_modelo,
                fila_destino=fila_destino,
                columnas_saltar=columnas_saltar
            )

            ws_publicar.cell(fila_destino, col_titulo).value = titulo
            ws_publicar.cell(fila_destino, col_fotos).value = imagenes
            ws_publicar.cell(fila_destino, col_descripcion).value = descripcion

            if col_caracteres is not None:
                ws_publicar.cell(fila_destino, col_caracteres).value = len(str(titulo))

        resumen.append({
            "categoria": categoria,
            "hoja_publicar": hoja_publicar_real,
            "estado": "OK",
            "titulos_generados": len(titulos),
            "fila_modelo_base_completar": fila_modelo,
            "fila_inicio_publicar": fila_inicio,
            "col_titulo": col_titulo,
            "col_fotos": col_fotos,
            "col_descripcion": col_descripcion,
            "links_imagenes": contar_links_imagenes(imagenes),
            "caracteres_descripcion": len(descripcion),
        })

    if errores:
        detalle = "\n".join(
            f"- {e.get('categoria', '')}: {e.get('error', '')}"
            for e in errores
        )

        raise ValueError(
            "El proceso encontró errores. No se genera archivo para evitar una carga incompleta.\n\n"
            + detalle
        )

    if not resumen:
        raise ValueError("No se procesó ninguna categoría.")

    salida = BytesIO()
    wb_publicar.save(salida)
    salida.seek(0)

    return salida, resumen


def mostrar_herramienta_agente_publicador():
    st.title("Agente Publicador")
    st.caption("Completa el archivo Publicar usando el AGENTE PUBLICADOR BASE.")

    st.warning(
        "Esta herramienta es independiente de Integraly. "
        "No cruza precios ni stock. Solo completa títulos, fotos y descripciones en Publicar."
    )

    st.info(
        "El AGENTE debe tener: BASE_COMPLETAR, BASE_TITULOS, BASE_DESCRIPCIONES, "
        "BASE_IMAGENES y CONFIG_PUBLICADOR."
    )

    col1, col2 = st.columns(2)

    with col1:
        archivo_agente = st.file_uploader(
            "Subí AGENTE PUBLICADOR BASE",
            type=["xlsx"],
            key="agente_publicador_uploader"
        )

    with col2:
        archivo_publicar = st.file_uploader(
            "Subí archivo Publicar",
            type=["xlsx"],
            key="publicar_uploader"
        )

    if st.button("Procesar Agente Publicador", type="primary"):

        if archivo_agente is None or archivo_publicar is None:
            st.error("Tenés que subir el AGENTE PUBLICADOR y el archivo Publicar.")

        else:
            try:
                with st.spinner("Procesando Publicar..."):
                    salida, resumen = procesar_agente_publicador(
                        archivo_agente.getvalue(),
                        archivo_publicar.getvalue()
                    )

                st.success("Archivo Publicar actualizado correctamente.")

                st.subheader("Resumen")
                st.dataframe(pd.DataFrame(resumen), use_container_width=True)

                st.download_button(
                    label="Descargar Publicar actualizado",
                    data=salida,
                    file_name=archivo_publicar.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error("El proceso falló.")
                st.exception(e)


# ============================================================
# SIDEBAR / ROUTER PRINCIPAL
# ============================================================

st.sidebar.title("ML Toolkit")

herramienta = st.sidebar.selectbox(
    "Elegí herramienta",
    [
        "Actualizar Integraly",
        "Agente Publicador",
    ]
)

if herramienta == "Actualizar Integraly":
    mostrar_herramienta_integraly()

elif herramienta == "Agente Publicador":
    mostrar_herramienta_agente_publicador()
